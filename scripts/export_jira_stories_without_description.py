"""Export Jira stories created during the last year without description to Excel.

Usage:
    python scripts/export_jira_stories_without_description.py
    python scripts/export_jira_stories_without_description.py --output stories_without_description.xlsx

Required environment variables:
    JIRA_URL
    JIRA_API_TOKEN

Optional environment variables:
    JIRA_USERNAME or JIRA_EMAIL  - required only for basic auth
    JIRA_AUTH_MODE              - token (default) or basic
    JIRA_VERIFY_SSL             - 1/true/yes to verify SSL certificates
"""

from __future__ import annotations

import argparse
import importlib.util
import os
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Iterator, Optional, Sequence, Tuple

script_dir = Path(__file__).resolve().parent
parent_dir = script_dir.parent
if str(parent_dir) not in sys.path:
    sys.path.insert(0, str(parent_dir))
if str(script_dir) not in sys.path:
    sys.path.insert(0, str(script_dir))

DEFAULT_PROJECT_KEYS: tuple[str, ...] = (
    "HRM",
    "HRC",
    "PERFREVIEW",
    "HRPASSIST",
    "SFILE",
    "NEUROUI",
    "SEARCHCS",
    "SBRPPL",
)
DEFAULT_ISSUE_TYPE = "Story"
DEFAULT_DAYS_BACK = 365
JIRA_SEARCH_FIELDS: tuple[str, ...] = (
    "summary",
    "description",
    "status",
    "assignee",
    "reporter",
    "created",
    "updated",
    "priority",
    "labels",
)
EXCEL_HEADERS: tuple[str, ...] = (
    "Key",
    "Summary",
    "Status",
    "Assignee",
    "Reporter",
    "Created",
    "Updated",
    "Priority",
    "Labels",
    "URL",
)


@dataclass(frozen=True)
class JiraSettings:
    """Connection settings for Jira REST API."""

    url: str
    token: str
    login: str
    auth_mode: str
    verify_ssl: bool


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments."""

    parser = argparse.ArgumentParser(
        description=(
            "Выгружает Story без описания, созданные за последний год, "
            "в Excel: один лист на Jira project."
        )
    )
    parser.add_argument(
        "--output",
        default=f"jira_stories_without_description_{datetime.now().strftime('%Y%m%d')}.xlsx",
        help="Путь к итоговому .xlsx файлу.",
    )
    parser.add_argument(
        "--projects",
        default=",".join(DEFAULT_PROJECT_KEYS),
        help="Список Jira project key через запятую.",
    )
    parser.add_argument(
        "--days",
        type=int,
        default=DEFAULT_DAYS_BACK,
        help="Сколько дней назад учитывать по created.",
    )
    parser.add_argument(
        "--issue-type",
        default=DEFAULT_ISSUE_TYPE,
        help="Тип задачи в Jira. По умолчанию Story.",
    )
    parser.add_argument(
        "--max-per-project",
        type=int,
        default=0,
        help="Ограничение задач на проект. 0 означает без ограничения.",
    )
    parser.add_argument(
        "--auth-mode",
        choices=("token", "basic"),
        default=None,
        help="Способ авторизации: token=Bearer token, basic=login+token.",
    )
    return parser.parse_args()


def parse_project_keys(raw_projects: str) -> tuple[str, ...]:
    """Return normalized project keys from a comma-separated string."""

    projects = tuple(project.strip().upper() for project in raw_projects.split(",") if project.strip())
    if not projects:
        raise ValueError("Список проектов пустой.")
    return projects


def env_bool(name: str, default: bool = False) -> bool:
    """Read boolean value from environment."""

    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}


def load_legacy_config() -> dict[str, Any]:
    """Load optional legacy config dict used by existing repository scripts."""

    script_dir = Path(__file__).resolve().parent
    candidate_paths = (
        script_dir.parent / "config.py",
        script_dir / "config.py",
    )
    for config_path in candidate_paths:
        if not config_path.exists():
            continue
        spec = importlib.util.spec_from_file_location("_jira_export_legacy_config", config_path)
        if spec is None or spec.loader is None:
            continue
        module = importlib.util.module_from_spec(spec)
        try:
            spec.loader.exec_module(module)
        except Exception:
            continue
        config = getattr(module, "config", None)
        if isinstance(config, dict):
            return config
    return {}


def load_jira_settings(auth_mode_arg: Optional[str] = None) -> JiraSettings:
    """Load Jira connection settings from .env, env vars, or legacy config."""

    try:
        from dotenv import load_dotenv
    except ImportError:
        load_dotenv = None

    if load_dotenv is not None:
        load_dotenv()

    legacy_config = load_legacy_config()
    legacy_jira = legacy_config.get("jira", {}) if legacy_config else {}
    legacy_options = legacy_jira.get("options", {}) if isinstance(legacy_jira, dict) else {}

    jira_url = (
        os.getenv("JIRA_URL", "").strip().rstrip("/")
        or str(legacy_jira.get("url", "")).strip().rstrip("/")
        or str(legacy_options.get("server", "")).strip().rstrip("/")
    )
    token = os.getenv("JIRA_API_TOKEN", "").strip() or str(legacy_jira.get("token", "")).strip()
    login = (os.getenv("JIRA_USERNAME", "") or os.getenv("JIRA_EMAIL", "")).strip()
    auth_mode = (auth_mode_arg or os.getenv("JIRA_AUTH_MODE", "token")).strip().lower()
    verify_ssl = env_bool("JIRA_VERIFY_SSL", default=False)

    missing = []
    if not jira_url:
        missing.append("JIRA_URL")
    if not token:
        missing.append("JIRA_API_TOKEN")
    if auth_mode == "basic" and not login:
        missing.append("JIRA_USERNAME or JIRA_EMAIL")
    if missing:
        raise RuntimeError(f"Не заданы обязательные параметры Jira: {', '.join(missing)}.")
    if auth_mode not in {"token", "basic"}:
        raise RuntimeError("JIRA_AUTH_MODE должен быть token или basic.")

    return JiraSettings(
        url=jira_url,
        token=token,
        login=login,
        auth_mode=auth_mode,
        verify_ssl=verify_ssl,
    )


def jql_value(value: str) -> str:
    """Return a safe JQL value representation."""

    if re.fullmatch(r"[A-Za-z][A-Za-z0-9_]*", value):
        return value
    escaped = value.replace("\\", "\\\\").replace('"', '\\"')
    return f'"{escaped}"'


def build_jql(project_key: str, issue_type: str, days_back: int) -> str:
    """Build JQL for stories without description in one project."""

    if days_back <= 0:
        raise ValueError("--days должен быть положительным числом.")
    return (
        f"project = {jql_value(project_key)} "
        f"AND issuetype = {jql_value(issue_type)} "
        f"AND description IS EMPTY "
        f"AND created >= -{days_back}d "
        "ORDER BY created DESC"
    )


def jira_auth(settings: JiraSettings) -> Tuple[dict[str, str], Optional[Tuple[str, str]]]:
    """Return headers and optional basic auth tuple for requests."""

    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    if settings.auth_mode == "basic":
        return headers, (settings.login, settings.token)
    headers["Authorization"] = f"Bearer {settings.token}"
    return headers, None


def request_jira_search(
    settings: JiraSettings,
    jql: str,
    start_at: int,
    max_results: int,
) -> dict[str, Any]:
    """Request one Jira search page."""

    try:
        import requests
        import urllib3
    except ImportError as exc:
        raise RuntimeError("Не установлены зависимости requests/urllib3. Выполните: pip install -r requirements.txt") from exc

    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    headers, auth = jira_auth(settings)
    response = requests.post(
        f"{settings.url}/rest/api/2/search",
        headers=headers,
        auth=auth,
        verify=settings.verify_ssl,
        timeout=60,
        json={
            "jql": jql,
            "startAt": start_at,
            "maxResults": max_results,
            "fields": list(JIRA_SEARCH_FIELDS),
        },
    )
    if response.status_code != 200:
        raise RuntimeError(
            f"Jira search failed: HTTP {response.status_code}; response: {response.text[:500]}"
        )
    data = response.json()
    if not isinstance(data, dict):
        raise RuntimeError("Jira search returned an unexpected response.")
    return data


def iter_project_issues(
    settings: JiraSettings,
    project_key: str,
    issue_type: str,
    days_back: int,
    max_per_project: int = 0,
    page_size: int = 100,
) -> Iterator[dict[str, Any]]:
    """Yield Jira issues for one project."""

    jql = build_jql(project_key=project_key, issue_type=issue_type, days_back=days_back)
    start_at = 0
    emitted = 0

    while True:
        remaining = max_per_project - emitted if max_per_project > 0 else page_size
        current_page_size = min(page_size, remaining) if max_per_project > 0 else page_size
        if current_page_size <= 0:
            break

        data = request_jira_search(
            settings=settings,
            jql=jql,
            start_at=start_at,
            max_results=current_page_size,
        )
        issues = data.get("issues", [])
        if not isinstance(issues, list) or not issues:
            break

        for issue in issues:
            if isinstance(issue, dict) and is_empty_description(issue.get("fields", {}).get("description")):
                yield issue
                emitted += 1
                if max_per_project > 0 and emitted >= max_per_project:
                    return

        start_at += len(issues)
        total = int(data.get("total", 0) or 0)
        if start_at >= total:
            break


def description_contains_text(value: Any) -> bool:
    """Return True if a Jira description value contains visible text."""

    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, dict):
        text = value.get("text")
        if isinstance(text, str) and text.strip():
            return True
        content = value.get("content")
        if content is not None:
            return description_contains_text(content)
        return any(
            description_contains_text(child)
            for key, child in value.items()
            if key not in {"type", "attrs", "marks"}
        )
    if isinstance(value, list):
        return any(description_contains_text(item) for item in value)
    return bool(str(value).strip())


def is_empty_description(value: Any) -> bool:
    """Return True if Jira description is empty or contains no visible text."""

    return not description_contains_text(value)


def user_display_name(value: Any) -> str:
    """Extract a display name from Jira user field."""

    if not isinstance(value, dict):
        return ""
    return str(value.get("displayName") or value.get("name") or value.get("emailAddress") or "").strip()


def named_value(value: Any) -> str:
    """Extract a readable name from Jira named field."""

    if not isinstance(value, dict):
        return ""
    return str(value.get("name") or value.get("value") or "").strip()


def issue_to_row(issue: dict[str, Any], jira_url: str) -> list[str]:
    """Convert Jira issue JSON to an Excel row."""

    fields = issue.get("fields", {})
    if not isinstance(fields, dict):
        fields = {}
    key = str(issue.get("key", "")).strip()
    labels = fields.get("labels") or []
    labels_text = ", ".join(str(label) for label in labels) if isinstance(labels, list) else str(labels)
    return [
        key,
        str(fields.get("summary") or "").strip(),
        named_value(fields.get("status")),
        user_display_name(fields.get("assignee")) or "Не назначен",
        user_display_name(fields.get("reporter")),
        str(fields.get("created") or "").strip(),
        str(fields.get("updated") or "").strip(),
        named_value(fields.get("priority")),
        labels_text,
        f"{jira_url}/browse/{key}" if key else "",
    ]


def sanitize_sheet_title(title: str, used_titles: set[str]) -> str:
    """Return an Excel-compatible unique sheet title."""

    cleaned = re.sub(r"[\[\]:*?/\\]", "_", title).strip() or "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    index = 2
    while candidate in used_titles:
        suffix = f"_{index}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        index += 1
    used_titles.add(candidate)
    return candidate


def autofit_columns(worksheet: Any, rows: Sequence[Sequence[str]]) -> None:
    """Set readable column widths."""

    for index, header in enumerate(EXCEL_HEADERS, start=1):
        values = [header]
        values.extend(row[index - 1] for row in rows if index - 1 < len(row))
        width = min(max(len(str(value)) for value in values) + 2, 70)
        column_letter = worksheet.cell(row=1, column=index).column_letter
        worksheet.column_dimensions[column_letter].width = width


def write_excel(output_path: Path, project_rows: dict[str, list[list[str]]]) -> None:
    """Write collected Jira issues to an Excel workbook."""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("Не установлена библиотека openpyxl. Выполните: pip install -r requirements.txt") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_sheet_titles: set[str] = set()
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for project_key, rows in project_rows.items():
        worksheet = workbook.create_sheet(sanitize_sheet_title(project_key, used_sheet_titles))
        worksheet.append(EXCEL_HEADERS)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            worksheet.append(row)
            key_cell = worksheet.cell(row=worksheet.max_row, column=1)
            url_cell = worksheet.cell(row=worksheet.max_row, column=10)
            if url_cell.value:
                key_cell.hyperlink = url_cell.value
                key_cell.style = "Hyperlink"
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        autofit_columns(worksheet, rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def collect_project_rows(
    settings: JiraSettings,
    project_keys: Iterable[str],
    issue_type: str,
    days_back: int,
    max_per_project: int,
) -> dict[str, list[list[str]]]:
    """Collect export rows grouped by project key."""

    result: dict[str, list[list[str]]] = {}
    for project_key in project_keys:
        print(f"Собираю {project_key}: {build_jql(project_key, issue_type, days_back)}")
        rows = [
            issue_to_row(issue, settings.url)
            for issue in iter_project_issues(
                settings=settings,
                project_key=project_key,
                issue_type=issue_type,
                days_back=days_back,
                max_per_project=max_per_project,
            )
        ]
        result[project_key] = rows
        print(f"  найдено: {len(rows)}")
    return result


def main() -> int:
    """Run export from CLI."""

    args = parse_args()
    try:
        project_keys = parse_project_keys(args.projects)
        settings = load_jira_settings(auth_mode_arg=args.auth_mode)
        project_rows = collect_project_rows(
            settings=settings,
            project_keys=project_keys,
            issue_type=args.issue_type,
            days_back=args.days,
            max_per_project=args.max_per_project,
        )
        output_path = Path(args.output).expanduser().resolve()
        write_excel(output_path, project_rows)
    except Exception as exc:
        print(f"Ошибка: {exc}", file=sys.stderr)
        return 1

    total = sum(len(rows) for rows in project_rows.values())
    print(f"Готово: {output_path}")
    print(f"Всего задач: {total}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
