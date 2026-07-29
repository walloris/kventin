#!/usr/bin/env python3
"""
Выгрузка статистики Jira-релизов за последние N календарных дней в Excel.

Логика источника повторяет scripts/dpm2.py:
  * проект HRPRELEASE;
  * тип задачи Release 2.0;
  * фактическая дата релиза из config['jira']['fields']
    ['prod_installed_date_id'];
  * КЭ из customfield_18300.

Story и Bug считаются только среди задач, связанных с релизом связью
``consist of`` / ``is part of``. Счёт ведётся по уникальным Jira-ключам.

Запросы выполняются последовательно. Скрипт ограничивает их частоту и общее
количество, не использует автоматические retry. Точный GET нужен только для
связей релиза: Jira Search может возвращать урезанный ``issuelinks``.
"""

from __future__ import annotations

import argparse
import importlib
import json
import logging
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence
from zoneinfo import ZoneInfo


DEFAULT_PROJECT = "HRPRELEASE"
DEFAULT_RELEASE_ISSUE_TYPE = "Release 2.0"
DEFAULT_CREATED_SINCE = "2025-09-01"
DEFAULT_PROD_DATE_FIELD_ID = "customfield_19400"
DEFAULT_RELEASE_KE_FIELD_ID = "customfield_18300"
DEFAULT_RELEASE_TYPE_FIELD_ID = "customfield_23500"
DEFAULT_RELEASE_KE_JQL_FIELD = "КЭ"
DEFAULT_INSTALLED_STATUSES = frozenset(("установлен на пром",))
DEFAULT_LINK_KEYWORDS = ("consist", "part")
DEFAULT_STORY_TYPES = frozenset(("story", "user story", "история"))
DEFAULT_BUG_TYPES = frozenset(("bug", "defect", "ошибка"))
DEFAULT_TIMEZONE = "Europe/Moscow"
DEFAULT_DAYS = 90
DEFAULT_KE_BATCH_SIZE = 160
DEFAULT_ISSUE_BATCH_SIZE = 100
DEFAULT_PAGE_SIZE = 500
DEFAULT_MIN_REQUEST_INTERVAL_SECONDS = 1.0
DEFAULT_MAX_JIRA_REQUESTS = 150
DEFAULT_MAX_RELEASE_DETAIL_REQUESTS = 90
DEFAULT_CONNECT_TIMEOUT_SECONDS = 10
DEFAULT_READ_TIMEOUT_SECONDS = 60

HEADERS = (
    "Дата релиза",
    "Тип релиза",
    "КЭ",
    "Кластер",
    "Количество сторей в релизе",
    "Количество багов в релизе",
)


@dataclass(frozen=True)
class ServiceInfo:
    service_id: str
    cluster: str
    service: str


@dataclass(frozen=True)
class ReleaseRow:
    release_date: date
    release_type: str
    service_ids: tuple[str, ...]
    clusters: tuple[str, ...]
    story_count: int
    bug_count: int

    def as_excel_row(self) -> list[Any]:
        return [
            self.release_date,
            self.release_type,
            ", ".join(self.service_ids) if self.service_ids else "—",
            "; ".join(self.clusters) if self.clusters else "Не найден в справочнике",
            self.story_count,
            self.bug_count,
        ]


@dataclass
class JiraRequestGuard:
    min_interval_seconds: float = DEFAULT_MIN_REQUEST_INTERVAL_SECONDS
    max_requests: int = DEFAULT_MAX_JIRA_REQUESTS
    requests_made: int = 0
    _last_request_started_at: float | None = None

    def __post_init__(self) -> None:
        if self.min_interval_seconds < 0:
            raise ValueError("Минимальный интервал Jira-запросов не может быть < 0.")
        if self.max_requests <= 0:
            raise ValueError("Лимит Jira-запросов должен быть больше 0.")

    @property
    def remaining_requests(self) -> int:
        return self.max_requests - self.requests_made

    def require_capacity(self, request_count: int, label: str) -> None:
        if request_count < 0:
            raise ValueError("Число резервируемых Jira-запросов не может быть < 0.")
        if request_count > self.remaining_requests:
            raise RuntimeError(
                "Остановлено защитой Jira до начала серии запросов: "
                f"для операции «{label}» нужно {request_count}, доступно "
                f"{self.remaining_requests} из общего лимита {self.max_requests}."
            )

    def before_request(self, label: str) -> None:
        self.require_capacity(1, label)

        now = time.monotonic()
        if self._last_request_started_at is not None:
            wait_seconds = (
                self.min_interval_seconds
                - (now - self._last_request_started_at)
            )
            if wait_seconds > 0:
                time.sleep(wait_seconds)

        self.requests_made += 1
        self._last_request_started_at = time.monotonic()
        logging.debug(
            "Jira request %s/%s: %s",
            self.requests_made,
            self.max_requests,
            label,
        )


def normalized(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def unique_preserving_order(values: Iterable[str]) -> tuple[str, ...]:
    result: list[str] = []
    seen: set[str] = set()
    for raw in values:
        value = str(raw or "").strip()
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return tuple(result)


def raw_value(value: Any) -> Any:
    """Convert jira-python resources to their JSON-like ``raw`` representation."""

    raw = getattr(value, "raw", None)
    return raw if raw is not None else value


def field_value(issue: Any, field_id: str) -> Any:
    if isinstance(issue, Mapping):
        fields = issue.get("fields")
        if isinstance(fields, Mapping):
            return fields.get(field_id)
        return issue.get(field_id)

    fields = getattr(issue, "fields", None)
    if fields is not None and hasattr(fields, field_id):
        return getattr(fields, field_id)

    raw_issue = raw_value(issue)
    if isinstance(raw_issue, Mapping):
        raw_fields = raw_issue.get("fields")
        if isinstance(raw_fields, Mapping):
            return raw_fields.get(field_id)
    return None


def has_issue_field(issue: Any, field_id: str) -> bool:
    if isinstance(issue, Mapping):
        fields = issue.get("fields")
        return isinstance(fields, Mapping) and field_id in fields

    fields = getattr(issue, "fields", None)
    if fields is not None and hasattr(fields, field_id):
        return True

    raw_issue = raw_value(issue)
    if isinstance(raw_issue, Mapping):
        raw_fields = raw_issue.get("fields")
        return isinstance(raw_fields, Mapping) and field_id in raw_fields
    return False


def issue_key(issue: Any) -> str:
    if isinstance(issue, Mapping):
        return str(issue.get("key") or "").strip().upper()
    return str(getattr(issue, "key", "") or "").strip().upper()


def named_text(value: Any) -> str:
    value = raw_value(value)
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float, bool)):
        return str(value)
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
        return ", ".join(
            item for item in (named_text(entry) for entry in value) if item
        )
    if isinstance(value, Mapping):
        for key in ("value", "name", "displayName", "key", "id"):
            text = named_text(value.get(key))
            if text:
                return text
    for attribute in ("value", "name", "displayName", "key", "id"):
        text = named_text(getattr(value, attribute, None))
        if text:
            return text
    return ""


def parse_jira_date(value: Any) -> date | None:
    text = named_text(value)
    if not text:
        return None
    match = re.search(r"\d{4}-\d{2}-\d{2}", text)
    if not match:
        return None
    try:
        return date.fromisoformat(match.group(0))
    except ValueError:
        return None


def parse_end_date(value: str | None, timezone_name: str) -> date:
    if value:
        try:
            return date.fromisoformat(value)
        except ValueError as exc:
            raise ValueError("--end-date должен иметь формат YYYY-MM-DD.") from exc
    try:
        return datetime.now(ZoneInfo(timezone_name)).date()
    except Exception as exc:
        raise ValueError(f"Неизвестная таймзона: {timezone_name}.") from exc


def period_bounds(end: date, days: int) -> tuple[date, date]:
    if days <= 0:
        raise ValueError("--days должен быть положительным числом.")
    return end - timedelta(days=days - 1), end


def load_service_catalog(path: Path) -> dict[str, ServiceInfo]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except FileNotFoundError as exc:
        raise FileNotFoundError(f"Не найден справочник КЭ: {path}") from exc
    except json.JSONDecodeError as exc:
        raise ValueError(f"Справочник {path} не является валидным JSON: {exc}") from exc

    if not isinstance(payload, list):
        raise ValueError("Корнем справочника КЭ должен быть JSON-массив.")

    result: dict[str, ServiceInfo] = {}
    for index, item in enumerate(payload, start=1):
        if not isinstance(item, Mapping):
            raise ValueError(f"Запись #{index} справочника не является объектом.")
        service_id = str(item.get("serviceId") or "").strip()
        cluster = str(item.get("cluster") or "").strip()
        service = str(item.get("service") or "").strip()
        if not service_id or not cluster:
            raise ValueError(
                f"В записи #{index} обязательны непустые serviceId и cluster."
            )
        previous = result.get(service_id)
        current = ServiceInfo(service_id, cluster, service)
        if previous is not None and previous != current:
            raise ValueError(f"Конфликтующие записи для serviceId={service_id}.")
        result[service_id] = current

    if not result:
        raise ValueError("Справочник КЭ пуст.")
    return result


def service_name_index(catalog: Mapping[str, ServiceInfo]) -> dict[str, str]:
    result: dict[str, str] = {}
    ambiguous: set[str] = set()
    for service_id, info in catalog.items():
        name = normalized(info.service)
        if not name:
            continue
        if name in result and result[name] != service_id:
            ambiguous.add(name)
        else:
            result[name] = service_id
    for name in ambiguous:
        result.pop(name, None)
    return result


def partial_catalog_service_ids(
    text: str,
    catalog: Mapping[str, ServiceInfo],
    by_name: Mapping[str, str],
) -> tuple[str, ...]:
    """
    Find КЭ by a partial Jira value.

    Jira may return values such as ``SmartProfile (2295205)`` or
    ``КЭ=2295205 / SmartProfile`` instead of a bare serviceId. Exact matches
    still have priority. A shortened value is accepted only when it points to
    exactly one catalog entry.
    """

    normalized_text = normalized(text)
    if not normalized_text:
        return ()

    service_ids_by_normalized = {
        normalized(service_id): service_id for service_id in catalog
    }
    exact_service_id = service_ids_by_normalized.get(normalized_text)
    if exact_service_id:
        return (exact_service_id,)

    contained_service_ids = sorted(
        (
            normalized_text.find(normalized_service_id),
            -len(normalized_service_id),
            service_id,
        )
        for normalized_service_id, service_id in service_ids_by_normalized.items()
        if normalized_service_id and normalized_service_id in normalized_text
    )
    if contained_service_ids:
        return unique_preserving_order(
            service_id for _, _, service_id in contained_service_ids
        )

    # Не матчим слишком короткие фрагменты: они почти гарантированно
    # неоднозначны среди сотен КЭ.
    if len(normalized_text) >= 4:
        shortened_id_matches = unique_preserving_order(
            service_id
            for normalized_service_id, service_id in service_ids_by_normalized.items()
            if normalized_text in normalized_service_id
        )
        if len(shortened_id_matches) == 1:
            return shortened_id_matches

    exact_name = by_name.get(normalized_text)
    if exact_name:
        return (exact_name,)

    contained_names = [
        (len(name), service_id)
        for name, service_id in by_name.items()
        if len(name) >= 4 and name in normalized_text
    ]
    if contained_names:
        longest_length = max(length for length, _ in contained_names)
        longest_matches = unique_preserving_order(
            service_id
            for length, service_id in contained_names
            if length == longest_length
        )
        if len(longest_matches) == 1:
            return longest_matches

    if len(normalized_text) >= 4:
        shortened_name_matches = unique_preserving_order(
            service_id
            for name, service_id in by_name.items()
            if normalized_text in name
        )
        if len(shortened_name_matches) == 1:
            return shortened_name_matches

    return ()


def extract_service_ids(
    raw_ke_value: Any,
    catalog: Mapping[str, ServiceInfo],
) -> tuple[str, ...]:
    by_name = service_name_index(catalog)
    found: list[str] = []

    def visit(value: Any) -> None:
        value = raw_value(value)
        if value is None:
            return
        if isinstance(value, Sequence) and not isinstance(
            value, (str, bytes, bytearray)
        ):
            for item in value:
                visit(item)
            return
        if isinstance(value, Mapping):
            candidates = []
            for candidate in (
                value.get("serviceId"),
                value.get("id"),
                value.get("smId"),
            ):
                text = str(candidate or "").strip()
                if not text:
                    continue
                candidates.append(text)
                matched_ids = partial_catalog_service_ids(text, catalog, by_name)
                if matched_ids:
                    found.extend(matched_ids)
                    return
            for key in ("value", "name"):
                if value.get(key) is not None:
                    previous_count = len(found)
                    visit(value.get(key))
                    if len(found) > previous_count:
                        return
            if candidates:
                found.append(candidates[0])
            return

        text = str(value).strip()
        if not text:
            return
        matched_ids = partial_catalog_service_ids(text, catalog, by_name)
        if matched_ids:
            found.extend(matched_ids)
            return
        if re.fullmatch(r"[A-Za-zА-Яа-я0-9_-]+", text):
            found.append(text)

    visit(raw_ke_value)
    return unique_preserving_order(found)


def issue_link_type_text(link: Any) -> str:
    raw_link = raw_value(link)
    if isinstance(raw_link, Mapping):
        raw_type = raw_value(raw_link.get("type"))
    else:
        raw_type = raw_value(getattr(link, "type", None))

    if isinstance(raw_type, Mapping):
        return " ".join(
            named_text(raw_type.get(key))
            for key in ("name", "inward", "outward")
            if named_text(raw_type.get(key))
        )
    return named_text(raw_type)


def linked_issue_key(link: Any, direction: str) -> str:
    raw_link = raw_value(link)
    key_name = f"{direction}Issue"
    if isinstance(raw_link, Mapping):
        linked = raw_value(raw_link.get(key_name))
    else:
        linked = raw_value(getattr(link, key_name, None))

    if isinstance(linked, Mapping):
        return str(linked.get("key") or "").strip().upper()
    return str(getattr(linked, "key", "") or "").strip().upper()


def release_linked_keys(
    release: Any,
    keywords: Sequence[str] = DEFAULT_LINK_KEYWORDS,
) -> tuple[str, ...]:
    links = field_value(release, "issuelinks") or []
    if not isinstance(links, Sequence) or isinstance(links, (str, bytes, bytearray)):
        return ()

    normalized_keywords = tuple(normalized(keyword) for keyword in keywords)

    def keyword_matches(link_type: str, keyword: str) -> bool:
        if keyword == "consist":
            return bool(
                re.search(r"\bconsist(?:s|ed|ing)?\b|\bcomposition\b", link_type)
            )
        if keyword == "part":
            return bool(re.search(r"\bpart\b", link_type))
        return bool(re.search(rf"\b{re.escape(keyword)}\b", link_type))

    keys: list[str] = []
    for link in links:
        link_type = normalized(issue_link_type_text(link))
        if not any(
            keyword_matches(link_type, keyword) for keyword in normalized_keywords
        ):
            continue
        for direction in ("outward", "inward"):
            key = linked_issue_key(link, direction)
            if key:
                keys.append(key)
    return unique_preserving_order(keys)


def release_type(release: Any, release_type_field_id: str) -> str:
    raw_type = normalized(named_text(field_value(release, release_type_field_id)))
    summary = normalized(named_text(field_value(release, "summary")))
    hotfix_markers = ("hotfix", "hot-fix", "хотфикс")
    return (
        "Хотфикс"
        if any(marker in raw_type or marker in summary for marker in hotfix_markers)
        else "Релиз"
    )


def is_installed_release(release: Any) -> bool:
    return (
        normalized(named_text(field_value(release, "status")))
        in DEFAULT_INSTALLED_STATUSES
    )


def jql_value(value: str) -> str:
    text = str(value).strip()
    if re.fullmatch(r"\d+", text):
        return text
    return '"' + text.replace("\\", "\\\\").replace('"', '\\"') + '"'


def jql_field(field: str) -> str:
    value = field.strip()
    if not value:
        raise ValueError("JQL-поле не может быть пустым.")
    custom_field_match = re.fullmatch(
        r"customfield_(\d+)",
        value,
        flags=re.IGNORECASE,
    )
    if custom_field_match:
        return f"cf[{custom_field_match.group(1)}]"
    if re.fullmatch(r"cf\[\d+]|[\wА-Яа-яЁё]+", value):
        return value
    return jql_value(value)


def batches(values: Sequence[str], size: int) -> Iterable[Sequence[str]]:
    if size <= 0:
        raise ValueError("Размер пакета должен быть положительным.")
    for start in range(0, len(values), size):
        yield values[start : start + size]


def build_release_jql(
    service_ids: Sequence[str],
    *,
    created_since: str,
    project: str = DEFAULT_PROJECT,
    issue_type: str = DEFAULT_RELEASE_ISSUE_TYPE,
    ke_jql_field: str = DEFAULT_RELEASE_KE_JQL_FIELD,
    prod_date_field: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    installed_only: bool = False,
) -> str:
    values = ", ".join(jql_value(value) for value in service_ids)
    if (start_date is None) != (end_date is None):
        raise ValueError("Для JQL-периода нужны одновременно start_date и end_date.")
    if start_date is not None and end_date is not None and start_date > end_date:
        raise ValueError("Начало JQL-периода не может быть позже конца.")
    if start_date is not None and not prod_date_field:
        raise ValueError("Для JQL-периода нужно поле даты релиза.")

    clauses = [
        f"project = {jql_value(project)}",
        f"{jql_field(ke_jql_field)} in ({values})",
        f"type = {jql_value(issue_type)}",
        f'created >= "{created_since}"',
    ]
    if start_date is not None and end_date is not None:
        date_field = jql_field(str(prod_date_field))
        clauses.extend(
            (
                f'{date_field} >= "{start_date.isoformat()}"',
                f'{date_field} <= "{end_date.isoformat()}"',
            )
        )
    if installed_only:
        clauses.append('status = "Установлен на ПРОМ"')
    return " AND ".join(clauses)


def search_all(
    jira: Any,
    jql: str,
    fields: Sequence[str],
    *,
    page_size: int,
    request_guard: JiraRequestGuard | None = None,
    request_label: str = "JQL search",
) -> list[Any]:
    result: list[Any] = []
    start_at = 0
    while True:
        if request_guard is not None:
            request_guard.before_request(
                f"{request_label}, startAt={start_at}, maxResults={page_size}"
            )
        page = jira.search_issues(
            jql,
            startAt=start_at,
            maxResults=page_size,
            fields=",".join(dict.fromkeys(fields)),
        )
        items = list(page)
        result.extend(items)
        if not items:
            break
        start_at += len(items)
        total = int(getattr(page, "total", 0) or 0)
        if total:
            if start_at >= total:
                break
        elif len(items) < page_size:
            break
    return result


def collect_release_candidates(
    jira: Any,
    service_ids: Sequence[str],
    *,
    created_since: str,
    ke_jql_field: str,
    prod_date_field: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    installed_only: bool = False,
    release_fields: Sequence[str],
    ke_batch_size: int,
    page_size: int,
    verbose: bool,
    request_guard: JiraRequestGuard | None = None,
) -> dict[str, Any]:
    result: dict[str, Any] = {}
    service_id_list = list(unique_preserving_order(service_ids))
    batch_list = list(batches(service_id_list, ke_batch_size))
    for index, service_batch in enumerate(batch_list, start=1):
        jql = build_release_jql(
            service_batch,
            created_since=created_since,
            ke_jql_field=ke_jql_field,
            prod_date_field=prod_date_field,
            start_date=start_date,
            end_date=end_date,
            installed_only=installed_only,
        )
        if verbose:
            logging.info(
                "Поиск релизов: пакет КЭ %s/%s (%s значений)",
                index,
                len(batch_list),
                len(service_batch),
            )
        for issue in search_all(
            jira,
            jql,
            release_fields,
            page_size=page_size,
            request_guard=request_guard,
            request_label=f"релизы, пакет КЭ {index}/{len(batch_list)}",
        ):
            key = issue_key(issue)
            if key:
                result[key] = issue
    return result


def fetch_release_details(
    jira: Any,
    release_keys: Sequence[str],
    fields: Sequence[str],
    *,
    verbose: bool,
    preloaded: Mapping[str, Any] | None = None,
    max_detail_requests: int = DEFAULT_MAX_RELEASE_DETAIL_REQUESTS,
    request_guard: JiraRequestGuard | None = None,
) -> dict[str, Any]:
    result: dict[str, Any] = {}
    preloaded = preloaded or {}
    release_key_list = list(unique_preserving_order(release_keys))
    if max_detail_requests < 0:
        raise ValueError("Лимит точных GET релизов не может быть < 0.")
    if len(release_key_list) > max_detail_requests:
        raise RuntimeError(
            "Остановлено защитой Jira до точного чтения состава: найдено "
            f"{len(release_key_list)} релизов, разрешено не более "
            f"{max_detail_requests}. Уменьшите период либо осознанно увеличьте "
            "--max-release-detail-requests."
        )
    if request_guard is not None:
        request_guard.require_capacity(
            len(release_key_list),
            f"точный состав {len(release_key_list)} релизов",
        )

    requested_fields = tuple(dict.fromkeys(fields))
    full_field_csv = ",".join(requested_fields)
    for index, key in enumerate(release_key_list, start=1):
        candidate = preloaded.get(key)
        # Если метаданные уже пришли из Search, точным GET забираем только
        # issuelinks. В корпоративной Jira именно это поле в Search усекается.
        field_csv = "issuelinks" if candidate is not None else full_field_csv
        if request_guard is not None:
            request_guard.before_request(
                f"точный состав релиза {key} ({index}/{len(release_key_list)})"
            )
        try:
            exact_release = jira.issue(key, fields=field_csv)
        except Exception as exc:
            raise RuntimeError(
                f"Не удалось точно получить состав релиза {key}: {exc}"
            ) from exc
        if (
            not has_issue_field(exact_release, "issuelinks")
            or field_value(exact_release, "issuelinks") is None
        ):
            raise RuntimeError(
                f"Точный GET релиза {key} не вернул fields.issuelinks; "
                "нельзя достоверно посчитать Story/Bug."
            )

        if candidate is None:
            result[key] = exact_release
        else:
            raw_candidate = raw_value(candidate)
            raw_fields = (
                raw_candidate.get("fields")
                if isinstance(raw_candidate, Mapping)
                else None
            )
            if isinstance(raw_fields, Mapping):
                merged_fields = dict(raw_fields)
            else:
                merged_fields = {
                    field: field_value(candidate, field)
                    for field in requested_fields
                    if field != "issuelinks" and has_issue_field(candidate, field)
                }
            merged_fields["issuelinks"] = field_value(
                exact_release,
                "issuelinks",
            )
            result[key] = {"key": key, "fields": merged_fields}

        if verbose and (
            index == 1
            or index % 10 == 0
            or index == len(release_key_list)
        ):
            logging.info(
                "Точно загружен состав релизов: %s/%s",
                index,
                len(release_key_list),
            )
    return result


def fetch_linked_issues(
    jira: Any,
    issue_keys: Sequence[str],
    *,
    issue_batch_size: int,
    page_size: int,
    verbose: bool,
    request_guard: JiraRequestGuard | None = None,
) -> dict[str, Any]:
    result: dict[str, Any] = {}
    unique_keys = list(unique_preserving_order(issue_keys))
    batch_list = list(batches(unique_keys, issue_batch_size))
    if request_guard is not None:
        request_guard.require_capacity(
            len(batch_list),
            f"{len(batch_list)} пакетных запросов задач состава",
        )
    for index, key_batch in enumerate(batch_list, start=1):
        key_jql = ", ".join(jql_value(key) for key in key_batch)
        jql = f"key in ({key_jql})"
        loaded_in_batch: set[str] = set()
        for issue in search_all(
            jira,
            jql,
            ("issuetype",),
            page_size=page_size,
            request_guard=request_guard,
            request_label=f"задачи состава, пакет {index}/{len(batch_list)}",
        ):
            key = issue_key(issue)
            if key:
                result[key] = issue
                loaded_in_batch.add(key)
        missing_in_batch = [
            key for key in key_batch if key not in loaded_in_batch
        ]
        if missing_in_batch:
            raise RuntimeError(
                "Jira не вернула все задачи из consist of при пакетном поиске; "
                "выгрузка остановлена без одиночных повторных запросов. "
                f"Не получено: {', '.join(missing_in_batch[:20])}."
            )
        without_issue_type = [
            key
            for key in key_batch
            if not has_issue_field(result[key], "issuetype")
        ]
        if without_issue_type:
            raise RuntimeError(
                "Jira не вернула fields.issuetype для задач состава: "
                f"{', '.join(without_issue_type[:20])}; нельзя достоверно "
                "посчитать Story/Bug."
            )
        if verbose:
            logging.info(
                "Загружен пакет задач состава %s/%s", index, len(batch_list)
            )
    return result


def issue_type_name(issue: Any) -> str:
    return normalized(named_text(field_value(issue, "issuetype")))


def make_release_rows(
    releases: Mapping[str, Any],
    linked_issues: Mapping[str, Any],
    linked_keys_by_release: Mapping[str, Sequence[str]],
    catalog: Mapping[str, ServiceInfo],
    *,
    prod_date_field_id: str,
    ke_field_id: str,
    release_type_field_id: str,
) -> tuple[list[ReleaseRow], set[str]]:
    rows: list[ReleaseRow] = []
    unknown_service_ids: set[str] = set()

    for release_key in sorted(releases):
        release = releases[release_key]
        installed = parse_jira_date(field_value(release, prod_date_field_id))
        if installed is None:
            continue

        service_ids = extract_service_ids(field_value(release, ke_field_id), catalog)
        clusters: list[str] = []
        for service_id in service_ids:
            info = catalog.get(service_id)
            if info is None:
                unknown_service_ids.add(service_id)
                clusters.append(f"Не найден: {service_id}")
            else:
                clusters.append(info.cluster)

        story_keys: set[str] = set()
        bug_keys: set[str] = set()
        for key in linked_keys_by_release.get(release_key, ()):
            issue = linked_issues.get(key)
            if issue is None:
                continue
            issue_type = issue_type_name(issue)
            if issue_type in DEFAULT_STORY_TYPES:
                story_keys.add(key)
            elif issue_type in DEFAULT_BUG_TYPES:
                bug_keys.add(key)

        rows.append(
            ReleaseRow(
                release_date=installed,
                release_type=release_type(release, release_type_field_id),
                service_ids=service_ids,
                clusters=unique_preserving_order(clusters),
                story_count=len(story_keys),
                bug_count=len(bug_keys),
            )
        )

    rows.sort(
        key=lambda row: (
            row.release_date,
            row.release_type == "Хотфикс",
            row.service_ids,
        ),
        reverse=True,
    )
    return rows, unknown_service_ids


def write_excel(rows: Sequence[ReleaseRow], output_path: Path) -> None:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError(
            "Не установлен openpyxl. Выполните: pip install openpyxl>=3.1"
        ) from exc

    output_path = output_path.expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(f".{output_path.stem}.tmp.xlsx")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Релизы"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"

    sheet.append(list(HEADERS))
    for row in rows:
        sheet.append(row.as_excel_row())

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_border = Border(bottom=Side(style="medium", color="17365D"))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 34

    hotfix_fill = PatternFill("solid", fgColor="FFF2CC")
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row=row_index, column=1).number_format = "dd.mm.yyyy"
        sheet.cell(row=row_index, column=1).alignment = Alignment(horizontal="center")
        sheet.cell(row=row_index, column=2).alignment = Alignment(horizontal="center")
        sheet.cell(row=row_index, column=3).number_format = "@"
        sheet.cell(row=row_index, column=3).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        sheet.cell(row=row_index, column=4).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        for column in (5, 6):
            sheet.cell(row=row_index, column=column).number_format = "0"
            sheet.cell(row=row_index, column=column).alignment = Alignment(
                horizontal="right"
            )
        if sheet.cell(row=row_index, column=2).value == "Хотфикс":
            sheet.cell(row=row_index, column=2).fill = hotfix_fill

    widths = {
        "A": 16,
        "B": 15,
        "C": 24,
        "D": 48,
        "E": 26,
        "F": 25,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    last_row = max(sheet.max_row, 1)
    sheet.auto_filter.ref = f"A1:F{last_row}"
    sheet.print_area = f"A1:F{last_row}"
    sheet.print_title_rows = "1:1"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.5
    sheet.page_margins.bottom = 0.5
    if rows:
        table = Table(displayName="ReleaseStatsTable", ref=f"A1:F{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    workbook.save(temporary_path)

    check = load_workbook(temporary_path, read_only=True, data_only=True)
    try:
        check_sheet = check["Релизы"]
        actual_headers = tuple(cell.value for cell in check_sheet[1])
        if actual_headers != HEADERS:
            raise RuntimeError(
                f"Проверка XLSX не пройдена: заголовки {actual_headers!r}."
            )
        if check_sheet.max_row != len(rows) + 1:
            raise RuntimeError(
                "Проверка XLSX не пройдена: число строк отличается от ожидаемого."
            )
        for values in check_sheet.iter_rows(min_row=2, values_only=True):
            if not isinstance(values[0], (date, datetime)):
                raise RuntimeError("Дата релиза записана в XLSX не как дата.")
            if (
                not isinstance(values[4], int)
                or isinstance(values[4], bool)
                or not isinstance(values[5], int)
                or isinstance(values[5], bool)
            ):
                raise RuntimeError("Счётчики Story/Bug записаны в XLSX не как числа.")
    finally:
        check.close()

    os.replace(temporary_path, output_path)


def locate_config_module() -> Any:
    script_dir = Path(__file__).resolve().parent
    candidates = (Path.cwd(), script_dir, script_dir.parent)
    for candidate in candidates:
        candidate_text = str(candidate)
        if candidate_text not in sys.path:
            sys.path.insert(0, candidate_text)
    try:
        module = importlib.import_module("config")
    except ImportError as exc:
        raise RuntimeError(
            "Не найден config.py. Положите скрипт в папку scripts проекта "
            "рядом с родительским config.py либо запускайте из корня проекта."
        ) from exc
    return module


def create_jira_client(*, verify_ssl: bool) -> tuple[Any, Mapping[str, Any]]:
    try:
        from jira import JIRA
    except ImportError as exc:
        raise RuntimeError(
            "Не установлен jira. Выполните: pip install jira>=3.5"
        ) from exc

    config_module = locate_config_module()
    nested_config = getattr(config_module, "config", {})
    if isinstance(nested_config, Mapping):
        nested_jira = nested_config.get("jira")
        jira_config = nested_jira if isinstance(nested_jira, Mapping) else {}
    else:
        jira_config = {}

    nested_options = jira_config.get("options")
    options = dict(nested_options) if isinstance(nested_options, Mapping) else {}
    server = str(
        jira_config.get("url")
        or options.get("server")
        or getattr(config_module, "JIRA_URL", "")
        or os.getenv("JIRA_URL", "")
        or ""
    ).strip()
    token = str(
        jira_config.get("token")
        or getattr(config_module, "JIRA_API_TOKEN", "")
        or os.getenv("JIRA_API_TOKEN", "")
        or ""
    ).strip()
    username = str(
        jira_config.get("username")
        or jira_config.get("email")
        or getattr(config_module, "JIRA_USERNAME", "")
        or getattr(config_module, "JIRA_EMAIL", "")
        or os.getenv("JIRA_USERNAME", "")
        or os.getenv("JIRA_EMAIL", "")
        or ""
    ).strip()
    auth_mode = str(
        jira_config.get("auth_mode")
        or getattr(config_module, "JIRA_AUTH_MODE", "")
        or os.getenv("JIRA_AUTH_MODE", "")
        or "token"
    ).strip().casefold()
    # В корпоративной Jira используется self-signed CA. Это повторяет
    # рабочее поведение release_checker.py; проверку можно включить CLI-флагом.
    options["verify"] = verify_ssl
    if server:
        options["server"] = server.rstrip("/")

    if not options.get("server"):
        raise RuntimeError(
            "В config.py не задан Jira URL: нужен "
            "config['jira']['options']['server'] или JIRA_URL."
        )
    if not token:
        raise RuntimeError(
            "В config.py не задан Jira token: нужен "
            "config['jira']['token'] или JIRA_API_TOKEN."
        )
    client_options = {
        "options": options,
        # Не делаем скрытый /serverInfo при создании клиента и не умножаем
        # нагрузку автоматическими retry. Любая сетевая ошибка завершает запуск.
        "get_server_info": False,
        "max_retries": 0,
        "timeout": (
            DEFAULT_CONNECT_TIMEOUT_SECONDS,
            DEFAULT_READ_TIMEOUT_SECONDS,
        ),
    }
    if auth_mode == "basic":
        if not username:
            raise RuntimeError("Для Jira basic auth нужен username/email.")
        client = JIRA(**client_options, basic_auth=(username, token))
    elif auth_mode == "token":
        client = JIRA(**client_options, token_auth=token)
    else:
        raise RuntimeError("JIRA_AUTH_MODE должен быть token или basic.")
    return client, jira_config


def parse_arguments(argv: Sequence[str] | None = None) -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Сформировать Excel-статистику Jira-релизов за последние N дней."
    )
    parser.add_argument(
        "--mapping",
        type=Path,
        default=script_dir / "service_clusters.json",
        help="Декодированный JSON со связкой serviceId → cluster.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Путь итогового XLSX. По умолчанию — в текущей папке.",
    )
    parser.add_argument(
        "--days",
        type=int,
        default=DEFAULT_DAYS,
        help=f"Размер календарного окна, по умолчанию {DEFAULT_DAYS}.",
    )
    parser.add_argument(
        "--end-date",
        default=None,
        help="Последняя дата окна YYYY-MM-DD; по умолчанию сегодня.",
    )
    parser.add_argument(
        "--timezone",
        default=DEFAULT_TIMEZONE,
        help=f"Таймзона для определения сегодня; по умолчанию {DEFAULT_TIMEZONE}.",
    )
    parser.add_argument(
        "--created-since",
        default=DEFAULT_CREATED_SINCE,
        help=(
            "Дополнительная нижняя граница created из JQL dpm2.py. Окно "
            "последних N дней также фильтруется в Jira по дате установки."
        ),
    )
    parser.add_argument(
        "--prod-date-field",
        default=None,
        help="Jira field id даты релиза; по умолчанию берётся из config.py.",
    )
    parser.add_argument(
        "--ke-field",
        default=DEFAULT_RELEASE_KE_FIELD_ID,
        help=f"Jira field id КЭ; по умолчанию {DEFAULT_RELEASE_KE_FIELD_ID}.",
    )
    parser.add_argument(
        "--ke-jql-field",
        default=DEFAULT_RELEASE_KE_JQL_FIELD,
        help=(
            "Имя или customfield_* поля КЭ в JQL; по умолчанию "
            f"{DEFAULT_RELEASE_KE_JQL_FIELD}."
        ),
    )
    parser.add_argument(
        "--release-type-field",
        default=DEFAULT_RELEASE_TYPE_FIELD_ID,
        help=(
            "Jira field id типа релиза; при пустом поле используется summary."
        ),
    )
    parser.add_argument(
        "--verify-ssl",
        action="store_true",
        help=(
            "Проверять TLS-сертификат Jira. По умолчанию отключено, как в "
            "release_checker.py для корпоративного self-signed CA."
        ),
    )
    parser.add_argument(
        "--include-non-installed",
        action="store_true",
        help=(
            "Включить релизы не в статусе «Установлен на ПРОМ». По умолчанию "
            "плановые и отменённые задачи исключаются."
        ),
    )
    parser.add_argument(
        "--min-request-interval",
        type=float,
        default=DEFAULT_MIN_REQUEST_INTERVAL_SECONDS,
        help=(
            "Минимальная пауза между Jira-запросами в секундах; "
            f"по умолчанию {DEFAULT_MIN_REQUEST_INTERVAL_SECONDS}."
        ),
    )
    parser.add_argument(
        "--max-jira-requests",
        type=int,
        default=DEFAULT_MAX_JIRA_REQUESTS,
        help=(
            "Жёсткий лимит Jira-запросов за один запуск; "
            f"по умолчанию {DEFAULT_MAX_JIRA_REQUESTS}."
        ),
    )
    parser.add_argument(
        "--max-release-detail-requests",
        type=int,
        default=DEFAULT_MAX_RELEASE_DETAIL_REQUESTS,
        help=(
            "Предельное число последовательных точных GET состава релизов; "
            f"по умолчанию {DEFAULT_MAX_RELEASE_DETAIL_REQUESTS}."
        ),
    )
    parser.add_argument("--verbose", action="store_true", help="Подробный лог.")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_arguments(argv)
    request_guard: JiraRequestGuard | None = None
    logging.basicConfig(
        level=logging.INFO if args.verbose else logging.WARNING,
        format="%(asctime)s %(levelname)s %(message)s",
    )

    try:
        end_date = parse_end_date(args.end_date, args.timezone)
        start_date, end_date = period_bounds(end_date, args.days)
        output_path = (
            args.output
            if args.output is not None
            else Path.cwd()
            / f"release_stats_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
        )
        request_guard = JiraRequestGuard(
            min_interval_seconds=args.min_request_interval,
            max_requests=args.max_jira_requests,
        )
        if args.max_release_detail_requests < 0:
            raise ValueError(
                "--max-release-detail-requests не может быть меньше 0."
            )

        catalog = load_service_catalog(args.mapping.expanduser().resolve())
        jira, jira_config = create_jira_client(verify_ssl=args.verify_ssl)
        fields_config = jira_config.get("fields")
        if not isinstance(fields_config, Mapping):
            fields_config = {}
        prod_date_field_id = (
            args.prod_date_field
            or str(fields_config.get("prod_installed_date_id") or "").strip()
            or DEFAULT_PROD_DATE_FIELD_ID
        )

        release_fields = (
            "summary",
            "status",
            prod_date_field_id,
            args.ke_field,
            args.release_type_field,
        )
        candidates = collect_release_candidates(
            jira,
            tuple(catalog),
            created_since=args.created_since,
            ke_jql_field=args.ke_jql_field,
            prod_date_field=prod_date_field_id,
            start_date=start_date,
            end_date=end_date,
            installed_only=not args.include_non_installed,
            release_fields=release_fields,
            ke_batch_size=DEFAULT_KE_BATCH_SIZE,
            page_size=DEFAULT_PAGE_SIZE,
            verbose=args.verbose,
            request_guard=request_guard,
        )

        candidate_dates = {
            key: parse_jira_date(field_value(release, prod_date_field_id))
            for key, release in candidates.items()
        }
        if candidates and not any(value is not None for value in candidate_dates.values()):
            raise RuntimeError(
                f"Ни у одного из {len(candidates)} релизов Jira не вернула "
                f"дату из поля {prod_date_field_id}. Проверьте --prod-date-field "
                "и config['jira']['fields']['prod_installed_date_id']."
            )
        date_matched_keys = {
            key
            for key, installed in candidate_dates.items()
            if installed is not None and start_date <= installed <= end_date
        }
        if (
            date_matched_keys
            and not args.include_non_installed
            and not any(
                named_text(field_value(candidates[key], "status"))
                for key in date_matched_keys
            )
        ):
            raise RuntimeError(
                "Jira не вернула status для релизов выбранного периода; "
                "нельзя применить фильтр «Установлен на ПРОМ»."
            )
        in_period_keys = sorted(
            key
            for key, release in candidates.items()
            if (
                key in date_matched_keys
                and (
                    args.include_non_installed
                    or is_installed_release(release)
                )
            )
        )
        logging.info(
            "Кандидатов релизов=%s; в периоде %s—%s=%s",
            len(candidates),
            start_date.isoformat(),
            end_date.isoformat(),
            len(in_period_keys),
        )

        detail_fields = (*release_fields, "issuelinks")
        full_releases = fetch_release_details(
            jira,
            in_period_keys,
            detail_fields,
            verbose=args.verbose,
            preloaded=candidates,
            max_detail_requests=args.max_release_detail_requests,
            request_guard=request_guard,
        )
        for key, release in full_releases.items():
            if parse_jira_date(field_value(release, prod_date_field_id)) is None:
                raise RuntimeError(
                    f"Данные релиза {key} не содержат корректную дату "
                    f"из поля {prod_date_field_id}."
                )
            if (
                not args.include_non_installed
                and not named_text(field_value(release, "status"))
            ):
                raise RuntimeError(
                    f"Данные релиза {key} не содержат status; "
                    "нельзя применить фильтр «Установлен на ПРОМ»."
                )
        releases = {
            key: release
            for key, release in full_releases.items()
            if (
                (installed := parse_jira_date(
                    field_value(release, prod_date_field_id)
                ))
                is not None
                and start_date <= installed <= end_date
                and (
                    args.include_non_installed
                    or is_installed_release(release)
                )
            )
        }
        linked_keys_by_release = {
            key: release_linked_keys(release)
            for key, release in releases.items()
        }
        all_linked_keys = unique_preserving_order(
            linked_key
            for key in releases
            for linked_key in linked_keys_by_release.get(key, ())
        )
        linked_issues = fetch_linked_issues(
            jira,
            all_linked_keys,
            issue_batch_size=DEFAULT_ISSUE_BATCH_SIZE,
            page_size=DEFAULT_PAGE_SIZE,
            verbose=args.verbose,
            request_guard=request_guard,
        )

        rows, unknown_service_ids = make_release_rows(
            releases,
            linked_issues,
            linked_keys_by_release,
            catalog,
            prod_date_field_id=prod_date_field_id,
            ke_field_id=args.ke_field,
            release_type_field_id=args.release_type_field,
        )
        write_excel(rows, output_path)

        release_count = sum(row.release_type == "Релиз" for row in rows)
        hotfix_count = sum(row.release_type == "Хотфикс" for row in rows)
        print(f"Период: {start_date:%d.%m.%Y}–{end_date:%d.%m.%Y}")
        print(
            f"Строк: {len(rows)} (релизов: {release_count}, "
            f"хотфиксов: {hotfix_count})"
        )
        if unknown_service_ids:
            print(
                "Предупреждение: КЭ без кластера: "
                + ", ".join(sorted(unknown_service_ids)),
                file=sys.stderr,
            )
        print(
            "Jira-запросов: "
            f"{request_guard.requests_made}/{request_guard.max_requests}; "
            f"интервал ≥ {request_guard.min_interval_seconds:g} с"
        )
        print(f"Excel: {Path(output_path).expanduser().resolve()}")
        return 0
    except Exception as exc:
        logging.exception("Ошибка выполнения") if args.verbose else None
        print(f"Ошибка: {exc}", file=sys.stderr)
        if request_guard is not None:
            print(
                "Jira-запросов до остановки: "
                f"{request_guard.requests_made}/{request_guard.max_requests}",
                file=sys.stderr,
            )
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
